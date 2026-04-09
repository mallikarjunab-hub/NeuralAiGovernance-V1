"""
/api/export — Export query results to Excel or PDF (Enhancement 3)

Endpoints:
  POST /api/export/excel   { question, data, answer }  → .xlsx download
  POST /api/export/pdf     { question, data, answer }  → .pdf download

Both endpoints accept pre-fetched data (the frontend passes the last query's
data rows), so no extra database round-trip is needed.
"""
import io, logging
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional

router = APIRouter(prefix="/api/export", tags=["Export"])
logger = logging.getLogger(__name__)


class ExportRequest(BaseModel):
    question: str
    answer:   Optional[str] = None
    data:     Optional[list] = None
    chart_type: Optional[str] = None


# ── Excel export ──────────────────────────────────────────────────────────────

@router.post("/excel")
async def export_excel(req: ExportRequest):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DSSS Data"

    HDR_FILL  = PatternFill("solid", fgColor="003580")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    META_FONT = Font(italic=True, color="555555", name="Calibri", size=10)
    TITLE_FONT= Font(bold=True, color="003580", name="Calibri", size=13)
    THIN      = Side(style="thin", color="C7D9F5")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    now_str = datetime.now().strftime("%d %b %Y %H:%M")

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = "DSSS Analytics Export — Department of Social Welfare, Goa"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Metadata rows
    ws.append(["Question:", req.question or ""])
    ws.append(["Generated:", now_str])
    ws.append(["Answer:", req.answer or ""])
    for row_idx in [2, 3, 4]:
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="003580", name="Calibri", size=10)
        ws.cell(row=row_idx, column=2).font = META_FONT

    ws.append([])  # blank separator

    if req.data:
        cols = list(req.data[0].keys())
        # Header row
        hdr_row = ws.max_row + 1
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=hdr_row, column=ci,
                           value=col.replace("_", " ").title())
            cell.fill   = HDR_FILL
            cell.font   = HDR_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[hdr_row].height = 20

        # Data rows
        for ri, row in enumerate(req.data):
            dr = ws.max_row + 1
            for ci, col in enumerate(cols, 1):
                val = row.get(col)
                # Try to cast numerics
                try:
                    val = float(val) if "." in str(val) else int(val)
                except (ValueError, TypeError):
                    pass
                cell = ws.cell(row=dr, column=ci, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
                if ri % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="EDF3FF")

        # Auto-fit column widths
        for ci, col in enumerate(cols, 1):
            max_len = max(
                len(str(col)),
                *[len(str(row.get(col, ""))) for row in req.data[:50]]
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)
    else:
        ws.append(["No data rows to export."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"dssy_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── PDF export ────────────────────────────────────────────────────────────────

@router.post("/pdf")
async def export_pdf(req: ExportRequest):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise HTTPException(500, "reportlab not installed — run: pip install reportlab")

    buf = io.BytesIO()
    page_size = landscape(A4) if (req.data and len(req.data[0]) > 5) else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    GOV_BLUE  = colors.HexColor("#003580")
    LIGHT_BLU = colors.HexColor("#EDF3FF")
    MID_BLUE  = colors.HexColor("#D6E4FF")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GovTitle", parent=styles["Heading1"],
        textColor=GOV_BLUE, fontSize=14, spaceAfter=4,
        fontName="Helvetica-Bold"
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        textColor=colors.HexColor("#444444"), fontSize=9, spaceAfter=2
    )
    answer_style = ParagraphStyle(
        "Answer", parent=styles["Normal"],
        textColor=colors.HexColor("#1e3a6e"), fontSize=10,
        spaceAfter=8, leading=14, borderPadding=(6, 6, 6, 6),
        backColor=LIGHT_BLU, borderColor=MID_BLUE, borderWidth=1,
        borderRadius=4
    )

    story = []

    story.append(Paragraph(
        "DSSS Analytics Report — Directorate of Social Welfare, Government of Goa",
        title_style
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=GOV_BLUE, spaceAfter=6))
    story.append(Paragraph(
        f"<b>Question:</b> {req.question or ''}",
        meta_style
    ))
    story.append(Paragraph(
        f"<b>Generated:</b> {datetime.now().strftime('%d %b %Y %H:%M')}",
        meta_style
    ))
    story.append(Spacer(1, 6))

    if req.answer:
        story.append(Paragraph(f"<b>AI Insight:</b> {req.answer}", answer_style))

    story.append(Spacer(1, 10))

    if req.data:
        cols = list(req.data[0].keys())
        header = [col.replace("_", " ").title() for col in cols]
        rows   = [[str(row.get(c, "")) for c in cols] for row in req.data[:500]]

        tbl_data = [header] + rows
        col_w    = [(page_size[0] - 3*cm) / len(cols)] * len(cols)

        tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            # Header
            ("BACKGROUND",   (0, 0), (-1, 0), GOV_BLUE),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 8),
            ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
            ("TOPPADDING",   (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING",(0, 0), (-1, 0), 6),
            # Data rows
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT_BLU]),
            ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
            ("TOPPADDING",   (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
            # Grid
            ("GRID",         (0, 0), (-1, -1), 0.4, MID_BLUE),
            ("LINEBELOW",    (0, 0), (-1, 0), 1.5, GOV_BLUE),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No data rows to export.", meta_style))

    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_BLUE))
    story.append(Paragraph(
        "Neural AI Governance — DSSS Analytics Platform | Confidential",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       textColor=colors.HexColor("#8faad4"), fontSize=7, alignment=1)
    ))

    doc.build(story)
    buf.seek(0)

    fname = f"dssy_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
